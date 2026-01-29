from flask import Flask, render_template, request, redirect, url_for, flash, session, Response
from datetime import datetime
from dotenv import load_dotenv
import os
import logging
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from db_supabase import get_db_connection_wrapper

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

if not app.secret_key:
    raise ValueError("請在 .env 檔案中設定 SECRET_KEY！")

# Error handler for better debugging
@app.errorhandler(500)
def internal_error(error):
    logger.error(f'Internal server error: {error}', exc_info=True)
    return render_template('error.html', error=str(error)), 500

# Custom filter to format datetime objects
@app.template_filter('format_datetime')
def format_datetime(value):
    """Format datetime object to yyyy-mm-dd hh:mm:ss"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return value

def get_db_connection():
    return get_db_connection_wrapper()

# 取得目前登入使用者
def get_current_user():
    if 'user_id' not in session:
        return None
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE id = %s', (session['user_id'],)).fetchone()
    conn.close()
    return user

# 判斷是否為管理員
def is_admin(user):
    return user['is_admin'] is True if user else False

# 判斷目前使用者是否有權編輯或刪除指定 bug
def can_edit_or_delete(bug, user):
    if not user:
        return False
    if is_admin(user):
        return True
    return bug['reported_by_user_id'] == user['id']

# 首頁 - 錯誤記錄列表（登入後才顯示記錄）
@app.route('/', methods=['GET'])
def index():
    user = get_current_user()
    
    if user:
        query = request.args.get('query', '')
        conn = get_db_connection()

        if query:
            bugs = conn.execute('''
                SELECT b.*, u.username as reporter_username 
                FROM bugs b LEFT JOIN users u ON b.reported_by_user_id = u.id
                WHERE b.bug_details ILIKE %s OR b.system ILIKE %s OR b.notes ILIKE %s
                ORDER BY b.report_date DESC
            ''', (f'%{query}%', f'%{query}%', f'%{query}%')).fetchall()
        else:
            bugs = conn.execute('''
                SELECT b.*, u.username as reporter_username 
                FROM bugs b LEFT JOIN users u ON b.reported_by_user_id = u.id
                ORDER BY b.report_date DESC
            ''').fetchall()
        conn.close()

        bugs_with_permission = []
        for bug in bugs:
            bug_dict = dict(bug)
            bug_dict['can_edit'] = can_edit_or_delete(bug, user)
            bugs_with_permission.append(bug_dict)
        
        return render_template('index.html',
                               bugs=bugs_with_permission,
                               query=query,
                               user=user,
                               show_list=True)
    else:
        return render_template('index.html',
                               bugs=[],
                               query='',
                               user=None,
                               show_list=False)

# 新增錯誤記錄（所有人皆可）
@app.route('/add', methods=['GET', 'POST'])
def add_bug():
    user = get_current_user()
    if request.method == 'POST':
        system = request.form['system'].strip()
        bug_details = request.form['bug_details'].strip()
        reported_by = request.form['reported_by'].strip()
        status = request.form['status']
        priority = request.form['priority']
        severity = request.form['severity']
        assigned_to = request.form.get('assigned_to', '').strip()
        notes = request.form.get('notes', '').strip()
        reported_by_user_id = user['id'] if user else None

        if status in ['已解決', '已關閉'] and not notes:
            flash('當狀態設為「已解決」或「已關閉」時，必須填寫備註說明解決方式或關閉原因！', 'error')
            return render_template('add.html', user=user)

        conn = get_db_connection()
        conn.execute('''
            INSERT INTO bugs 
            (report_date, system, bug_details, reported_by, status, priority, severity, assigned_to, notes, reported_by_user_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (datetime.now(), system, bug_details, reported_by, status, priority, severity, assigned_to, notes, reported_by_user_id))
        conn.commit()
        conn.close()
        flash('記錄新增成功！')
        return redirect(url_for('index'))

    return render_template('add.html', user=user)

# 編輯錯誤記錄（系統不可修改）
@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_bug(id):
    user = get_current_user()
    if not user:
        flash('請先登入才能編輯記錄！', 'error')
        return redirect(url_for('login'))

    conn = get_db_connection()
    bug = conn.execute('SELECT * FROM bugs WHERE id = %s', (id,)).fetchone()
    conn.close()

    if bug is None:
        flash('找不到該錯誤記錄！')
        return redirect(url_for('index'))

    if bug['status'] in ['已解決', '已關閉']:
        flash(f'此錯誤記錄已「{bug["status"]}」，無法再進行編輯！', 'error')
        return redirect(url_for('index'))

    if not can_edit_or_delete(bug, user):
        flash('您沒有權限編輯此記錄！', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        bug_details = request.form['bug_details'].strip()
        reported_by = request.form['reported_by'].strip()
        status = request.form['status']
        priority = request.form['priority']
        severity = request.form['severity']
        assigned_to = request.form.get('assigned_to', '').strip()
        notes = request.form.get('notes', '').strip()

        if status in ['已解決', '已關閉'] and not notes:
            flash('當狀態設為「已解決」或「已關閉」時，必須填寫備註！', 'error')
            return render_template('edit.html', bug=bug, user=user)

        resolution_date = bug['resolution_date']
        if bug['status'] not in ['已解決', '已關閉'] and status in ['已解決', '已關閉']:
            resolution_date = datetime.now()

        conn = get_db_connection()
        conn.execute('''
            UPDATE bugs 
            SET bug_details = %s, reported_by = %s, status = %s, priority = %s, severity = %s, 
                assigned_to = %s, notes = %s, resolution_date = %s
            WHERE id = %s
        ''', (bug_details, reported_by, status, priority, severity, assigned_to, notes, resolution_date, id))
        conn.commit()
        conn.close()
        flash('記錄更新成功！')
        return redirect(url_for('index'))

    return render_template('edit.html', bug=bug, user=user)

# 刪除錯誤記錄
@app.route('/delete/<int:id>', methods=['POST'])
def delete_bug(id):
    user = get_current_user()
    if not user:
        flash('請先登入才能刪除記錄！', 'error')
        return redirect(url_for('login'))

    conn = get_db_connection()
    bug = conn.execute('SELECT * FROM bugs WHERE id = %s', (id,)).fetchone()

    if bug and can_edit_or_delete(bug, user):
        conn.execute('DELETE FROM bugs WHERE id = %s', (id,))
        conn.commit()
        flash('錯誤記錄刪除成功！')
    else:
        flash('您沒有權限刪除此記錄！', 'error')

    conn.close()
    return redirect(url_for('index'))

# 使用者登入
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = %s', (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            flash('登入成功！', 'success')
            return redirect(url_for('index'))
        else:
            flash('使用者名稱或密碼錯誤！', 'error')

    return render_template('login.html')

# 使用者註冊
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            confirm_password = request.form.get('confirm_password', '')

            if not username or not password:
                flash('使用者名稱與密碼皆為必填！', 'error')
                return render_template('register.html')

            if password != confirm_password:
                flash('兩次輸入的密碼不一致！', 'error')
                return render_template('register.html')

            if len(password) < 6:
                flash('密碼長度至少需 6 個字元！', 'error')
                return render_template('register.html')

            conn = get_db_connection()
            try:
                existing_user = conn.execute('SELECT * FROM users WHERE username = %s', (username,)).fetchone()
                if existing_user:
                    flash('此使用者名稱已被使用，請選擇其他名稱！', 'error')
                    return render_template('register.html')

                password_hash = generate_password_hash(password)
                conn.execute('''
                    INSERT INTO users (username, password_hash, is_admin)
                    VALUES (%s, %s, FALSE)
                ''', (username, password_hash))
                conn.commit()
                
                logger.info(f"New user registered: {username}")
                flash('註冊成功！請登入使用。', 'success')
                return redirect(url_for('login'))
            finally:
                conn.close()
                
        except Exception as e:
            logger.error(f"Registration error for {username if 'username' in locals() else 'unknown'}: {str(e)}", exc_info=True)
            flash(f'註冊過程中發生錯誤: {str(e)}', 'error')
            return render_template('register.html')

    return render_template('register.html')

# 使用者登出
@app.route('/logout')
def logout():
    session.clear()
    flash('已成功登出。')
    return redirect(url_for('index'))

# 個人變更密碼
@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    user = get_current_user()
    if not user:
        flash('請先登入！', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if not check_password_hash(user['password_hash'], current_password):
            flash('目前密碼錯誤！', 'error')
            return render_template('change_password.html', user=user)

        if new_password != confirm_password:
            flash('兩次新密碼不一致！', 'error')
            return render_template('change_password.html', user=user)

        if len(new_password) < 6:
            flash('新密碼長度至少需 6 個字元！', 'error')
            return render_template('change_password.html', user=user)

        conn = get_db_connection()
        new_hash = generate_password_hash(new_password)
        conn.execute('UPDATE users SET password_hash = %s WHERE id = %s', (new_hash, user['id']))
        conn.commit()
        conn.close()

        flash('密碼變更成功！請重新登入。', 'success')
        session.clear()
        return redirect(url_for('login'))

    return render_template('change_password.html', user=user)

# admin 專用：變更任何使用者密碼
@app.route('/admin_change_password/<int:user_id>', methods=['GET', 'POST'])
def admin_change_password(user_id):
    current_user = get_current_user()
    if not current_user or current_user['username'] != 'admin':
        flash('僅限 admin 帳號使用此功能！', 'error')
        return redirect(url_for('index'))

    conn = get_db_connection()
    target_user = conn.execute('SELECT * FROM users WHERE id = %s', (user_id,)).fetchone()
    if target_user is None:
        flash('找不到該使用者！', 'error')
        conn.close()
        return redirect(url_for('index'))

    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if new_password != confirm_password:
            flash('兩次新密碼不一致！', 'error')
            conn.close()
            return render_template('admin_change_password.html', target_user=target_user)

        if len(new_password) < 6:
            flash('新密碼長度至少需 6 個字元！', 'error')
            conn.close()
            return render_template('admin_change_password.html', target_user=target_user)

        new_hash = generate_password_hash(new_password)
        conn.execute('UPDATE users SET password_hash = %s WHERE id = %s', (new_hash, user_id))
        conn.commit()
        conn.close()

        flash(f'已成功為使用者「{target_user["username"]}」變更密碼！', 'success')
        return redirect(url_for('index'))

    conn.close()
    return render_template('admin_change_password.html', target_user=target_user)

# 匯出 Excel 報表（僅登入使用者）
@app.route('/export_excel')
def export_excel():
    user = get_current_user()
    if not user:
        flash('請先登入才能匯出報表！', 'error')
        return redirect(url_for('login'))

    conn = get_db_connection()
    bugs = conn.execute('''
        SELECT b.id, b.report_date, b.system, b.bug_details, b.reported_by,
               b.status, b.priority, b.severity, b.assigned_to, b.resolution_date, b.notes,
               u.username as reporter_username
        FROM bugs b LEFT JOIN users u ON b.reported_by_user_id = u.id
        ORDER BY b.report_date DESC
    ''').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "錯誤追蹤報表"

    colors = {
        '開放中': 'FFFF99', '處理中': 'ADD8E6', '已解決': '90EE90', '已關閉': 'D3D3D3',
        '低': 'D3D3D3', '中': 'ADD8E6', '高': 'FFB6C1',
        '輕微': 'D3D3D3', '中': 'ADD8E6', '重大': 'FFFF99', '嚴重': 'FFB6C1',
    }

    headers = ['ID', '報告日期', '系統', '錯誤細節', '報告者', '報告者帳號', '狀態', '優先級', '嚴重程度', '指派給', '解決日期', '備註']
    ws.append(headers)

    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for bug in bugs:
        # Format timestamps for display
        report_date_str = bug['report_date'].strftime('%Y-%m-%d %H:%M:%S') if bug['report_date'] else ''
        resolution_date_str = bug['resolution_date'].strftime('%Y-%m-%d %H:%M:%S') if bug['resolution_date'] else ''
        
        row = [
            bug['id'], report_date_str, bug['system'], bug['bug_details'],
            bug['reported_by'], bug['reporter_username'] or '（未登入使用者）',
            bug['status'], bug['priority'], bug['severity'],
            bug['assigned_to'] or '', resolution_date_str, bug['notes'] or ''
        ]
        ws.append(row)

        row_idx = ws.max_row
        ws.cell(row=row_idx, column=7).fill = PatternFill(start_color=colors.get(bug['status'], 'FFFFFF'), end_color=colors.get(bug['status'], 'FFFFFF'), fill_type='solid')
        ws.cell(row=row_idx, column=8).fill = PatternFill(start_color=colors.get(bug['priority'], 'FFFFFF'), end_color=colors.get(bug['priority'], 'FFFFFF'), fill_type='solid')
        ws.cell(row=row_idx, column=9).fill = PatternFill(start_color=colors.get(bug['severity'], 'FFFFFF'), end_color=colors.get(bug['severity'], 'FFFFFF'), fill_type='solid')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": f"attachment;filename=bug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )

if __name__ == '__main__':
    app.run(debug=True)